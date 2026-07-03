"""
analytics_tickets.py - Excel exporter for Analytics Team Jira tickets.
Reads analytics_tickets_data.json (produced by analytics_tickets.ps1)
and writes a formatted Excel file. No network calls are made here.
"""

import os
import json
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("Run: pip install openpyxl")

_DIR        = os.path.dirname(os.path.abspath(__file__))
_JSON_PATH  = os.path.join(_DIR, "analytics_tickets_data.json")
_JIRA_BASE  = "https://yourcompany.atlassian.net/browse"


def write_excel(records, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analytics Tickets"

    header_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    center       = Alignment(horizontal="center", vertical="center")
    left         = Alignment(horizontal="left",   vertical="center")
    alt_fill     = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    link_font    = Font(color="1F4E79", underline="single", size=10)

    # Status cell fill colors keyed to Jira status names
    status_colors = {
        "Open":        "E3F2FD",
        "In Progress": "FFF8E1",
        "Waiting for Support":   "F3E5F5",
        "Waiting for Approval":  "FFF3E0",
        "Resolved":    "E8F5E9",
        "Closed":      "ECEFF1",
        "Done":        "ECEFF1",
    }

    headers    = ["Ticket Number", "Title", "Status", "Date Requested", "Last Updated", "Requested By", "Assigned To"]
    col_widths = [18, 52, 24, 20, 20, 28, 28]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    for row_idx, rec in enumerate(records, 2):
        fill = alt_fill if row_idx % 2 == 0 else None

        # Col 1 - ticket key as a clickable hyperlink
        key  = rec.get("ticket_key", "")
        cell = ws.cell(row=row_idx, column=1, value=key)
        if key:
            cell.hyperlink = f"{_JIRA_BASE}/{key}"
            cell.font      = link_font
        cell.alignment = center
        if fill:
            cell.fill = fill

        # Col 2 - title / summary
        cell           = ws.cell(row=row_idx, column=2, value=rec.get("title", ""))
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if fill:
            cell.fill = fill

        # Col 3 - status with soft background color per status
        status = rec.get("status", "")
        cell   = ws.cell(row=row_idx, column=3, value=status)
        cell.alignment = center
        hex_color = status_colors.get(status)
        if hex_color:
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        elif fill:
            cell.fill = fill

        # Col 4 - date requested
        cell           = ws.cell(row=row_idx, column=4, value=rec.get("date_requested", ""))
        cell.alignment = center
        if fill:
            cell.fill = fill

        # Col 5 - last updated
        cell           = ws.cell(row=row_idx, column=5, value=rec.get("last_updated", ""))
        cell.alignment = center
        if fill:
            cell.fill = fill

        # Col 6 - requested by
        cell           = ws.cell(row=row_idx, column=6, value=rec.get("requested_by", ""))
        cell.alignment = left
        if fill:
            cell.fill = fill

        # Col 7 - assigned to
        cell           = ws.cell(row=row_idx, column=7, value=rec.get("assigned_to", ""))
        cell.alignment = left
        if fill:
            cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(records) + 1}"

    wb.save(output_path)
    print(f"Excel saved: {output_path} ({len(records)} rows)")


if __name__ == "__main__":
    import sys

    if not os.path.exists(_JSON_PATH):
        print(f"ERROR: {_JSON_PATH} not found. Run analytics_tickets.ps1 first.")
        sys.exit(1)

    with open(_JSON_PATH, "r", encoding="utf-8-sig") as f:
        raw = f.read().strip()

    if not raw or raw == "null":
        print("No new tickets since last run. Exiting without generating Excel.")
        sys.exit(0)

    records = json.loads(raw)

    if isinstance(records, dict):
        records = [records]

    if not records:
        print("No new tickets since last run. Exiting without generating Excel.")
        sys.exit(0)

    stamp       = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = os.path.join(_DIR, f"analytics_tickets_{stamp}.xlsx")
    write_excel(records, output_path)
    print(f"Done! {len(records)} tickets exported.")
